"""Generate 4 sample Excel files for Fraud Guard Pro demo.
Each file has 2 sheets:
  Sheet 1 (Sao kê): Human-readable with VND amounts + transaction time
  Sheet 2 (Dữ liệu kỹ thuật): Time (seconds), Amount (USD = VND/25000), V1-V28
Both sheets are fully synchronized.
"""
import random, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

random.seed(42)
OUT_DIR = os.path.join(os.path.dirname(__file__), "sample_data")
os.makedirs(OUT_DIR, exist_ok=True)

VND_TO_USD = 25_000  # Tỉ giá quy đổi

HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="2563EB")
META_FONT = Font(bold=True, size=12)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
FRAUD_FILL = PatternFill("solid", fgColor="FEE2E2")


def rand_v_normal():
    return [round(random.uniform(-0.5, 0.5), 4) for _ in range(28)]


def rand_v_fraud():
    v = [round(random.uniform(-1.5, 1.5), 4) for _ in range(28)]
    v[0] = round(random.uniform(-5.5, -3.5), 4)   # V1
    v[2] = round(random.uniform(-6.0, -4.0), 4)   # V3
    v[3] = round(random.uniform(3.5, 5.5), 4)     # V4
    v[6] = round(random.uniform(-4.5, -2.5), 4)   # V7
    v[9] = round(random.uniform(-6.5, -4.5), 4)   # V10
    v[10] = round(random.uniform(3.0, 5.0), 4)    # V11
    v[11] = round(random.uniform(-8.2, -6.0), 4)  # V12
    v[13] = round(random.uniform(-10.5, -7.5), 4) # V14
    v[15] = round(random.uniform(-5.0, -3.0), 4)  # V16
    v[16] = round(random.uniform(-14.0, -10.0), 4)# V17
    v[17] = round(random.uniform(-3.8, -1.8), 4)  # V18
    return v


def time_str(hour, minute):
    """Format hh:mm."""
    return f"{hour:02d}:{minute:02d}"


def time_to_seconds(hour, minute):
    """Convert hh:mm to seconds from midnight."""
    return hour * 3600 + minute * 60


def build_workbook(customer, bank, account, transactions):
    """
    transactions: list of dict with keys:
        date, desc, tx_type, amount_vnd, balance_vnd, note,
        is_fraud, hour, minute
    """
    wb = Workbook()

    # ── Sheet 1: Sao kê giao dịch ──
    ws1 = wb.active
    ws1.title = "Sao ke giao dich"
    ws1.merge_cells("A1:H1")
    ws1["A1"] = f"Khách hàng: {customer}"
    ws1["A1"].font = META_FONT
    ws1.merge_cells("A2:H2")
    ws1["A2"] = f"Ngân hàng: {bank}"
    ws1["A2"].font = Font(bold=True, size=11)
    ws1.merge_cells("A3:H3")
    ws1["A3"] = f"Số tài khoản: {account}"
    ws1["A3"].font = Font(size=11)

    headers1 = ["Ngày GD", "Giờ GD", "Nội dung/Diễn giải", "Tài khoản",
                "Loại GD", "Số tiền (VNĐ)", "Số dư (VNĐ)", "Ghi chú"]
    for c, h in enumerate(headers1, 1):
        cell = ws1.cell(row=5, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    for i, tx in enumerate(transactions):
        row = 6 + i
        vals = [
            tx["date"],
            time_str(tx["hour"], tx["minute"]),
            tx["desc"],
            account,
            tx["tx_type"],
            tx["amount_vnd"],
            tx["balance_vnd"],
            tx["note"],
        ]
        for c, v in enumerate(vals, 1):
            cell = ws1.cell(row=row, column=c, value=v)
            cell.border = THIN_BORDER
            if tx["is_fraud"]:
                cell.fill = FRAUD_FILL

    ws1.column_dimensions["A"].width = 14
    ws1.column_dimensions["B"].width = 10
    ws1.column_dimensions["C"].width = 35
    ws1.column_dimensions["D"].width = 18
    ws1.column_dimensions["E"].width = 10
    ws1.column_dimensions["F"].width = 18
    ws1.column_dimensions["G"].width = 18
    ws1.column_dimensions["H"].width = 15

    # ── Sheet 2: Dữ liệu kỹ thuật ──
    ws2 = wb.create_sheet("Du lieu ky thuat")
    headers2 = ["Row", "Time", "Amount"] + [f"V{i}" for i in range(1, 29)]
    for c, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = PatternFill("solid", fgColor="1E293B")
        cell.alignment = Alignment(horizontal="center")

    for i, tx in enumerate(transactions):
        row = 2 + i
        is_fraud = tx["is_fraud"]
        v_vals = rand_v_fraud() if is_fraud else rand_v_normal()

        # Time = seconds from midnight of the transaction
        time_sec = time_to_seconds(tx["hour"], tx["minute"])
        # Amount = VND converted to USD
        amount_usd = round(abs(tx["amount_vnd"]) / VND_TO_USD, 2)

        ws2.cell(row=row, column=1, value=i + 1)
        ws2.cell(row=row, column=2, value=time_sec)
        ws2.cell(row=row, column=3, value=amount_usd)
        for j, v in enumerate(v_vals):
            ws2.cell(row=row, column=4 + j, value=v)

    return wb


def make_tx(date, hour, minute, desc, tx_type, amount_vnd, note, is_fraud, balance_ref):
    """Helper to build a transaction dict."""
    balance_ref[0] += amount_vnd
    return {
        "date": date,
        "hour": hour,
        "minute": minute,
        "desc": desc,
        "tx_type": tx_type,
        "amount_vnd": amount_vnd,
        "balance_vnd": balance_ref[0],
        "note": note,
        "is_fraud": is_fraud,
    }


# ══════════════════════════════════════════════════
# File 1: Nguyễn Văn An - VCB - Bình thường
# ══════════════════════════════════════════════════
bal = [5_000_000]
txs1 = [
    make_tx("01/05/2026", 8, 30, "Nhận lương T4/2026 - Cty ABC", "Thu", 20_000_000, "", False, bal),
    make_tx("02/05/2026", 9, 15, "Thanh toán tiền nhà tháng 5", "Chi", -5_000_000, "", False, bal),
    make_tx("03/05/2026", 7, 45, "Grab - Đi làm Q1 đến Q7", "Chi", -35_000, "", False, bal),
    make_tx("03/05/2026", 12, 10, "Cơm trưa - Quán Bà Hai", "Chi", -45_000, "", False, bal),
    make_tx("04/05/2026", 8, 5, "Highland Coffee - Cà phê sáng", "Chi", -55_000, "", False, bal),
    make_tx("05/05/2026", 14, 30, "Shopee - Mua áo thun Uniqlo", "Chi", -350_000, "", False, bal),
    make_tx("06/05/2026", 7, 50, "Grab - Đi làm", "Chi", -32_000, "", False, bal),
    make_tx("07/05/2026", 10, 0, "Tiền điện EVN T4", "Chi", -385_000, "", False, bal),
    make_tx("08/05/2026", 10, 20, "Tiền nước Sawaco T4", "Chi", -120_000, "", False, bal),
    make_tx("09/05/2026", 9, 0, "Internet VNPT tháng 5", "Chi", -200_000, "", False, bal),
    make_tx("10/05/2026", 18, 30, "Grab - Đi ăn tối", "Chi", -28_000, "", False, bal),
    make_tx("10/05/2026", 19, 15, "Nhà hàng Hải Sản Biển Đông", "Chi", -450_000, "", False, bal),
    make_tx("12/05/2026", 11, 0, "Chuyển khoản - Trả tiền anh Tuấn", "Chi", -500_000, "", False, bal),
    make_tx("13/05/2026", 12, 30, "GrabFood - Bún bò Huế", "Chi", -65_000, "", False, bal),
    make_tx("14/05/2026", 15, 45, "Circle K - Nước ngọt, snack", "Chi", -42_000, "", False, bal),
    make_tx("15/05/2026", 17, 10, "Xăng xe - Petrolimex Q7", "Chi", -150_000, "", False, bal),
    make_tx("16/05/2026", 20, 0, "Shopee - Sạc dự phòng Anker", "Chi", -280_000, "", False, bal),
    make_tx("18/05/2026", 9, 30, "Tiền điện thoại Viettel T5", "Chi", -100_000, "", False, bal),
    make_tx("20/05/2026", 10, 15, "CK cho mẹ - Tiền thuốc", "Chi", -1_000_000, "", False, bal),
    make_tx("25/05/2026", 14, 0, "Bảo hiểm nhân thọ Manulife", "Chi", -800_000, "", False, bal),
]

wb1 = build_workbook("Nguyễn Văn An", "Vietcombank (VCB)", "001-8xx-xxx-789", txs1)
wb1.save(os.path.join(OUT_DIR, "KH_NguyenVanAn_VCB.xlsx"))


# ══════════════════════════════════════════════════
# File 2: Trần Thị Bình - BIDV - Bình thường
# ══════════════════════════════════════════════════
bal = [3_000_000]
txs2 = [
    make_tx("01/05/2026", 8, 0, "Nhận lương giáo viên T4", "Thu", 15_000_000, "", False, bal),
    make_tx("02/05/2026", 9, 30, "Tiền học phí con - Trường Lê Quý Đôn", "Chi", -3_500_000, "", False, bal),
    make_tx("03/05/2026", 7, 15, "Chợ Bến Thành - Mua thực phẩm", "Chi", -180_000, "", False, bal),
    make_tx("04/05/2026", 10, 0, "Tiền điện EVN T4", "Chi", -420_000, "", False, bal),
    make_tx("05/05/2026", 6, 45, "Chợ - Rau củ, thịt cá", "Chi", -95_000, "", False, bal),
    make_tx("06/05/2026", 14, 20, "Nhà thuốc Pharmacity", "Chi", -250_000, "", False, bal),
    make_tx("07/05/2026", 16, 30, "Xăng xe Honda Lead", "Chi", -120_000, "", False, bal),
    make_tx("08/05/2026", 10, 10, "Tiền nước T4", "Chi", -85_000, "", False, bal),
    make_tx("09/05/2026", 7, 0, "Chợ - Mua đồ nấu ăn", "Chi", -110_000, "", False, bal),
    make_tx("10/05/2026", 9, 45, "Bảo hiểm y tế gia đình", "Chi", -500_000, "", False, bal),
    make_tx("11/05/2026", 15, 0, "Sách giáo khoa cho con", "Chi", -320_000, "", False, bal),
    make_tx("13/05/2026", 7, 30, "Chợ Tân Định - Thực phẩm tuần", "Chi", -200_000, "", False, bal),
    make_tx("15/05/2026", 9, 0, "Tiền điện thoại Mobifone", "Chi", -79_000, "", False, bal),
    make_tx("17/05/2026", 14, 0, "Grab đi họp phụ huynh", "Chi", -45_000, "", False, bal),
    make_tx("19/05/2026", 6, 50, "Chợ - Đồ ăn cuối tuần", "Chi", -165_000, "", False, bal),
    make_tx("22/05/2026", 10, 30, "CK cho bà ngoại - Tiền thuốc", "Chi", -1_200_000, "", False, bal),
    make_tx("25/05/2026", 9, 15, "Internet FPT tháng 5", "Chi", -180_000, "", False, bal),
    make_tx("28/05/2026", 20, 45, "Tiki - Mua nồi cơm điện", "Chi", -890_000, "", False, bal),
]

wb2 = build_workbook("Trần Thị Bình", "BIDV", "210-0xx-xxx-456", txs2)
wb2.save(os.path.join(OUT_DIR, "KH_TranThiBinh_BIDV.xlsx"))


# ══════════════════════════════════════════════════
# File 3: Lê Hoàng Cường - TCB - Gian lận (cá cược xen kẽ)
# Fraud transactions happen at odd hours with betting-related descriptions
# ══════════════════════════════════════════════════
bal = [8_000_000]
txs3 = [
    make_tx("01/05/2026", 8, 30, "Nhận lương T4 - Cty XYZ Tech", "Thu", 18_000_000, "", False, bal),
    make_tx("02/05/2026", 7, 40, "Grab - Đi làm Thủ Đức", "Chi", -42_000, "", False, bal),
    make_tx("02/05/2026", 12, 15, "Cơm văn phòng - Canteen", "Chi", -35_000, "", False, bal),
    make_tx("03/05/2026", 21, 30, "Shopee - Tai nghe Bluetooth", "Chi", -290_000, "", False, bal),
    make_tx("04/05/2026", 15, 10, "Circle K - Nước uống", "Chi", -25_000, "", False, bal),
    # ⚠️ Fraud: Nạp tiền cá cược
    make_tx("05/05/2026", 23, 45, "NAP W88 ID7729456", "Chi", -2_000_000, "⚠️", True, bal),
    make_tx("05/05/2026", 10, 0, "Tiền điện T4 - EVN", "Chi", -310_000, "", False, bal),
    make_tx("06/05/2026", 12, 40, "GrabFood - Pizza 4P's", "Chi", -185_000, "", False, bal),
    # ⚠️ Fraud: CK nhà cái
    make_tx("07/05/2026", 1, 15, "CK LUCK88 xz7k9m2", "Chi", -3_000_000, "⚠️", True, bal),
    make_tx("08/05/2026", 17, 20, "Xăng xe - Shell Q2", "Chi", -135_000, "", False, bal),
    make_tx("09/05/2026", 18, 0, "Grab - Về quê cuối tuần", "Chi", -55_000, "", False, bal),
    # ⚠️ Fraud: Nạp tiền cược lớn
    make_tx("10/05/2026", 2, 30, "TOPUP 12BET DP55667", "Chi", -5_000_000, "⚠️", True, bal),
    make_tx("10/05/2026", 14, 50, "Cà phê The Coffee House", "Chi", -49_000, "", False, bal),
    # ⚠️ Fraud: Rút tiền thắng cược
    make_tx("12/05/2026", 3, 10, "RUT SBOBET 2.5TR ID884", "Thu", 2_500_000, "⚠️", True, bal),
    make_tx("13/05/2026", 12, 0, "Bún chả Hà Nội - Ăn trưa", "Chi", -55_000, "", False, bal),
    make_tx("14/05/2026", 9, 30, "Tiền điện thoại Viettel", "Chi", -100_000, "", False, bal),
    # ⚠️ Fraud: Rút tiền
    make_tx("15/05/2026", 4, 5, "CASHOUT FB88 uid_3391", "Thu", 1_800_000, "⚠️", True, bal),
    make_tx("16/05/2026", 20, 15, "Shopee - Quần áo hè", "Chi", -450_000, "", False, bal),
    # ⚠️ Fraud: CK đêm khuya
    make_tx("17/05/2026", 2, 47, "ck 02:47AM ho tro ky thuat abc123", "Chi", -1_500_000, "⚠️", True, bal),
    make_tx("18/05/2026", 8, 15, "Highland Coffee - Làm việc", "Chi", -65_000, "", False, bal),
    # ⚠️ Fraud: Thanh toán cá cược
    make_tx("19/05/2026", 22, 55, "thanh toan don hang 99bet ship cod", "Chi", -800_000, "⚠️", True, bal),
    make_tx("20/05/2026", 11, 0, "CK cho em gái - Sinh nhật", "Chi", -500_000, "", False, bal),
    make_tx("22/05/2026", 19, 30, "Grab - Đi ăn tối Q1", "Chi", -38_000, "", False, bal),
    make_tx("25/05/2026", 9, 0, "Internet VNPT T5", "Chi", -200_000, "", False, bal),
]

wb3 = build_workbook("Lê Hoàng Cường", "Techcombank (TCB)", "190-1xx-xxx-321", txs3)
wb3.save(os.path.join(OUT_DIR, "KH_LeHoangCuong_TCB.xlsx"))


# ══════════════════════════════════════════════════
# File 4: Phạm Minh Đức - MB - Gian lận (thẻ bị đánh cắp)
# Fraud transactions: sudden burst at 3AM-4AM = stolen card
# ══════════════════════════════════════════════════
bal = [12_000_000]
txs4 = [
    make_tx("01/05/2026", 8, 30, "Nhận lương T4 - Cty Logistics VN", "Thu", 22_000_000, "", False, bal),
    make_tx("02/05/2026", 7, 50, "Grab - Đi làm Q5", "Chi", -38_000, "", False, bal),
    make_tx("03/05/2026", 12, 0, "Cơm trưa - Quán Phở 24", "Chi", -75_000, "", False, bal),
    make_tx("04/05/2026", 19, 20, "Shopee - Balo laptop", "Chi", -420_000, "", False, bal),
    make_tx("05/05/2026", 10, 0, "Tiền điện EVN T4", "Chi", -380_000, "", False, bal),
    make_tx("06/05/2026", 7, 30, "Circle K - Ăn sáng", "Chi", -35_000, "", False, bal),
    make_tx("07/05/2026", 17, 45, "Xăng xe Petrolimex", "Chi", -160_000, "", False, bal),
    make_tx("08/05/2026", 12, 30, "GrabFood - Gà rán Popeyes", "Chi", -125_000, "", False, bal),
    make_tx("09/05/2026", 11, 15, "Chuyển khoản - Trả tiền bạn", "Chi", -300_000, "", False, bal),
    make_tx("10/05/2026", 10, 20, "Tiền nước Sawaco", "Chi", -95_000, "", False, bal),
    # ⚠️ Thẻ bị đánh cắp — burst 03:12AM - 03:18AM
    make_tx("12/05/2026", 3, 12, "ATM withdrawal BANGKOK 03:12AM", "Chi", -15_000_000, "⚠️", True, bal),
    make_tx("12/05/2026", 3, 15, "MOMO qr_scan unknown 03:15AM", "Chi", -8_000_000, "⚠️", True, bal),
    make_tx("12/05/2026", 3, 18, "POS LUXURY STORE HK 03:18AM", "Chi", -12_500_000, "⚠️", True, bal),
    make_tx("12/05/2026", 3, 25, "WIRE TRF offshore acc#xr99k2", "Chi", -20_000_000, "⚠️", True, bal),
    make_tx("13/05/2026", 4, 1, "Online purchase CRYPTO EX 04:01AM", "Chi", -9_800_000, "⚠️", True, bal),
    # Sau khi phát hiện, khóa thẻ, mở lại
    make_tx("15/05/2026", 14, 0, "Cà phê Highlands - Làm việc", "Chi", -55_000, "", False, bal),
    make_tx("16/05/2026", 8, 45, "Grab - Đi họp", "Chi", -42_000, "", False, bal),
    make_tx("17/05/2026", 12, 15, "Cơm trưa văn phòng", "Chi", -50_000, "", False, bal),
    make_tx("18/05/2026", 21, 0, "Shopee - Ốp lưng điện thoại", "Chi", -89_000, "", False, bal),
    make_tx("20/05/2026", 9, 30, "Tiền điện thoại Vinaphone", "Chi", -100_000, "", False, bal),
    make_tx("25/05/2026", 9, 15, "Internet FPT T5", "Chi", -220_000, "", False, bal),
]

wb4 = build_workbook("Phạm Minh Đức", "MB Bank", "068-0xx-xxx-654", txs4)
wb4.save(os.path.join(OUT_DIR, "KH_PhamMinhDuc_MB.xlsx"))

print("✅ Đã tạo 4 file Excel trong:", OUT_DIR)
for f in sorted(os.listdir(OUT_DIR)):
    if f.endswith(".xlsx"):
        path = os.path.join(OUT_DIR, f)
        print(f"  📄 {f} ({os.path.getsize(path):,} bytes)")
