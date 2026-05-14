"""
app.py — FastAPI web server for Credit Card Fraud Detection.

Features:
    • Multi-model prediction (RF, LR, KNN, SVM across sampling strategies)
    • Dual AI chatbot (Gemini + Groq) with provider switching
    • Transaction history with PostgreSQL (SQLite fallback)
    • Excel batch import with format validation
    • Neo-Brutalism UI served from static/
"""

from __future__ import annotations

import json
import logging
import os
import sqlite3
import uuid
import tempfile
import io

import psycopg2
import psycopg2.extras
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional

import google.generativeai as genai
import httpx
import joblib
import numpy as np
import pandas as pd
from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException, UploadFile, File, Form
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles
from groq import Groq
from pydantic import BaseModel
import openpyxl

# ──────────────────────────────────────────────────
# Configuration
# ──────────────────────────────────────────────────

load_dotenv()

BASE_DIR = Path(__file__).resolve().parent
# Check inside BASE_DIR first (for Docker: /app/models), then parent (for local dev: Machine Learning/models)
MODELS_DIR = BASE_DIR / "models" if (BASE_DIR / "models").exists() else BASE_DIR.parent / "models"
DB_PATH = BASE_DIR / "fraud_history.db"
SCALE_COLS = ["Amount", "Time"]  # Must match src/config.py

# PostgreSQL configuration (from .env)
PG_CONFIG = {
    "dbname": os.getenv("PG_DBNAME", "fraud_guard"),
    "user": os.getenv("PG_USER", "postgres"),
    "password": os.getenv("PG_PASSWORD", ""),
    "host": os.getenv("PG_HOST", "localhost"),
    "port": os.getenv("PG_PORT", "5432"),
}
DB_TYPE = "postgresql"  # Will auto-fallback to "sqlite" if connection fails

GEMINI_API_KEY = os.getenv("GEMINI_API_KEY", "")
GROQ_API_KEY = os.getenv("GROQ_API_KEY", "")

logging.basicConfig(level=logging.INFO, format="%(asctime)s │ %(name)s │ %(message)s")
logger = logging.getLogger("webapp")

# ──────────────────────────────────────────────────
# Model Registry — 4 production models (SMOTE + Tuned Params)
# ──────────────────────────────────────────────────
# Chiến lược triển khai:
#   - 4 model SMOTE + Optuna Tuned Params = sản phẩm cuối cùng cho dự đoán
#   - Các model khác (baseline, undersample) vẫn nằm trong models/ để phục vụ
#     notebook so sánh / báo cáo, nhưng KHÔNG hiển thị trên giao diện web.

WEBAPP_MODELS = {
    "random_forest_smote": {
        "display_name": "Random Forest",
        "description": "Ensemble 200 cây quyết định, tuned bởi Optuna (F1=0.9999)",
    },
    "logistic_regression_smote": {
        "display_name": "Logistic Regression",
        "description": "Hồi quy logistic với ElasticNet, tuned bởi Optuna (F1=0.9524)",
    },
    "knn_smote": {
        "display_name": "KNN",
        "description": "K=3, khoảng cách Manhattan, tuned bởi Optuna (F1=0.9964)",
    },
    "svm_smote": {
        "display_name": "SVM",
        "description": "Kernel RBF, C=28.48, tuned bởi Optuna (F1=0.9988)",
    },
}


def discover_models() -> Dict[str, Any]:
    """Load the 4 production models (SMOTE + Tuned Params) for the web UI.

    Only models listed in WEBAPP_MODELS are exposed to the frontend.
    Other .pkl files in models/ (baseline, undersample) are kept for
    notebook comparisons / reports but NOT shown on the web interface.
    """
    available = {}
    scaler_path = MODELS_DIR / "robust_scaler.pkl"
    if not scaler_path.exists():
        logger.warning("Scaler not found at %s", scaler_path)
        return available

    for model_id, meta in WEBAPP_MODELS.items():
        pkl_file = MODELS_DIR / f"{model_id}.pkl"
        if pkl_file.exists():
            available[model_id] = {
                "file": str(pkl_file),
                "display_name": meta["display_name"],
                "description": meta["description"],
            }
            logger.info("Registered model: %s → %s", model_id, meta["display_name"])
        else:
            logger.warning("Model file not found: %s", pkl_file)

    return available


MODEL_REGISTRY = discover_models()
LOADED_MODELS: Dict[str, Any] = {}
SCALER = None


def load_scaler():
    global SCALER
    scaler_path = MODELS_DIR / "robust_scaler.pkl"
    if scaler_path.exists():
        SCALER = joblib.load(scaler_path)
        logger.info("Loaded scaler from %s", scaler_path)


def get_model(model_id: str):
    """Lazy-load and cache a model."""
    if model_id not in MODEL_REGISTRY:
        raise ValueError(f"Unknown model: {model_id}")

    if model_id not in LOADED_MODELS:
        import warnings
        path = MODEL_REGISTRY[model_id]["file"]
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            loaded = joblib.load(path)

        # Patch for sklearn version mismatch:
        # Models saved with sklearn>=1.8 may lack 'multi_class' attribute
        # which is required by sklearn<=1.6 for predict_proba()
        if hasattr(loaded, 'predict_proba') and not hasattr(loaded, 'multi_class'):
            loaded.multi_class = 'auto'

        LOADED_MODELS[model_id] = loaded
        logger.info("Loaded model: %s", model_id)

    return LOADED_MODELS[model_id]


# Load scaler on startup
load_scaler()

# ──────────────────────────────────────────────────
# Database — PostgreSQL with SQLite fallback
# ──────────────────────────────────────────────────


def get_db():
    """Get a database connection. Tries PostgreSQL first, falls back to SQLite."""
    global DB_TYPE
    if DB_TYPE == "postgresql":
        try:
            conn = psycopg2.connect(**PG_CONFIG)
            conn.autocommit = False
            return conn
        except Exception as e:
            logger.warning("PostgreSQL connection failed: %s — falling back to SQLite", e)
            DB_TYPE = "sqlite"

    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    """Create the fraud_history table if it doesn't exist."""
    conn = get_db()
    cur = conn.cursor()

    if DB_TYPE == "postgresql":
        cur.execute("""
            CREATE TABLE IF NOT EXISTS fraud_history (
                id SERIAL PRIMARY KEY,
                timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                model_used TEXT,
                time_val FLOAT,
                amount FLOAT,
                v_features JSONB,
                prediction INTEGER,
                probability FLOAT,
                verdict TEXT,
                batch_id TEXT,
                customer_name TEXT,
                bank_name TEXT,
                tx_description TEXT
            )
        """)
        # Add columns if they don't exist (migration for existing DBs)
        for col, ctype in [("batch_id", "TEXT"), ("customer_name", "TEXT"), ("bank_name", "TEXT"), ("tx_description", "TEXT")]:
            try:
                cur.execute(f"ALTER TABLE fraud_history ADD COLUMN {col} {ctype}")
            except Exception:
                conn.rollback()
                conn = get_db()
                cur = conn.cursor()
    else:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS fraud_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
                model_used TEXT,
                time_val FLOAT,
                amount FLOAT,
                v_features TEXT,
                prediction INTEGER,
                probability FLOAT,
                verdict TEXT,
                batch_id TEXT,
                customer_name TEXT,
                bank_name TEXT,
                tx_description TEXT
            )
        """)

    conn.commit()
    cur.close()
    conn.close()
    logger.info("Database initialized [%s] %s",
                DB_TYPE,
                PG_CONFIG['dbname'] if DB_TYPE == 'postgresql' else DB_PATH)


init_db()

# ──────────────────────────────────────────────────
# AI Providers
# ──────────────────────────────────────────────────

SYSTEM_PROMPT_CHAT = """Bạn là chuyên gia bảo mật tài chính cấp cao của hệ thống Fraud Guard Pro.
Dữ liệu bạn đang phân tích dựa trên bộ dữ liệu giao dịch thẻ tín dụng của Châu Âu với các đặc trưng PCA (V1-V28).
Hãy trả lời ngắn gọn, chuyên nghiệp, tập trung vào khía cạnh an ninh tài chính bằng tiếng Việt."""

SYSTEM_PROMPT_ANALYZE = """Bạn là Giám đốc Phân tích Rủi ro (CRO) của Fraud Guard Pro.
Bạn nổi tiếng với khả năng phân tích dữ liệu database cực kỳ chi tiết, khoa học và không bao giờ tóm tắt quá mức.
Hãy trả lời bằng tiếng Việt."""


async def call_gemini(prompt: str, system_msg: str) -> str:
    """Call Google Gemini API."""
    try:
        genai.configure(api_key=GEMINI_API_KEY)
        model = genai.GenerativeModel(
            "gemini-2.0-flash",
            system_instruction=system_msg,
        )
        response = model.generate_content(prompt)
        return response.text
    except Exception as e:
        logger.error("Gemini error: %s", e)
        return f"Lỗi Gemini: {str(e)}"


async def call_groq(prompt: str, system_msg: str) -> str:
    """Call Groq API."""
    try:
        client = Groq(api_key=GROQ_API_KEY)
        response = client.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[
                {"role": "system", "content": system_msg},
                {"role": "user", "content": prompt},
            ],
            max_tokens=2048,
        )
        return response.choices[0].message.content
    except Exception as e:
        logger.error("Groq error: %s", e)
        return f"Lỗi Groq: {str(e)}"


async def call_ai(prompt: str, system_msg: str, provider: str = "gemini") -> str:
    """Route to the selected AI provider with automatic fallback."""
    try:
        if provider == "groq":
            return await call_groq(prompt, system_msg)
        else:
            return await call_gemini(prompt, system_msg)
    except Exception:
        # Fallback to the other provider
        fallback = "groq" if provider == "gemini" else "gemini"
        logger.warning("Provider %s failed, falling back to %s", provider, fallback)
        if fallback == "groq":
            return await call_groq(prompt, system_msg)
        else:
            return await call_gemini(prompt, system_msg)


# ──────────────────────────────────────────────────
# Database helpers
# ──────────────────────────────────────────────────


def _ph() -> str:
    """Return the SQL placeholder for the current DB type."""
    return "%s" if DB_TYPE == "postgresql" else "?"


def get_history_context() -> str:
    """Build a text summary from recent history for AI context."""
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(
            "SELECT amount, verdict, probability, timestamp, model_used "
            "FROM fraud_history WHERE batch_id IS NULL ORDER BY timestamp DESC LIMIT 20"
        )
        rows = cur.fetchall()

        cur.execute(
            "SELECT AVG(amount), COUNT(*) FROM fraud_history WHERE verdict = 'An toàn' AND batch_id IS NULL"
        )
        safe_stats = cur.fetchone()

        cur.execute(
            "SELECT COUNT(*) FROM fraud_history WHERE verdict = 'Gian lận' AND batch_id IS NULL"
        )
        fraud_count = cur.fetchone()[0]

        cur.close()
        conn.close()

        avg_safe = safe_stats[0] if safe_stats[0] else 0
        total_safe = safe_stats[1] if safe_stats[1] else 0

        ctx = "TỔNG QUAN HỆ THỐNG (Database):\n"
        ctx += f"- Tổng giao dịch an toàn: {total_safe}\n"
        ctx += f"- Số tiền trung bình an toàn: {avg_safe:.2f}$\n"
        ctx += f"- Tổng vụ gian lận đã chặn: {fraud_count}\n\n"
        ctx += "GIAO DỊCH GẦN ĐÂY:\n"
        for r in rows[:10]:
            ctx += f"- {r['timestamp']}: {r['amount']}$ [{r['model_used']}] → {r['verdict']} ({r['probability']*100:.1f}%)\n"

        return ctx
    except Exception as e:
        return f"DB context không khả dụng: {e}"


# ──────────────────────────────────────────────────
# Pydantic Models
# ──────────────────────────────────────────────────

class Transaction(BaseModel):
    Time: float
    V1: float; V2: float; V3: float; V4: float; V5: float
    V6: float; V7: float; V8: float; V9: float; V10: float
    V11: float; V12: float; V13: float; V14: float; V15: float
    V16: float; V17: float; V18: float; V19: float; V20: float
    V21: float; V22: float; V23: float; V24: float; V25: float
    V26: float; V27: float; V28: float
    Amount: float


class PredictRequest(BaseModel):
    transaction: Transaction
    model_id: str = "random_forest_smote"


class ChatMessage(BaseModel):
    message: str
    provider: str = "gemini"
    context: Optional[dict] = None


class AnalysisRequest(BaseModel):
    data: dict
    prediction: int
    probability: float
    model_used: str = ""
    provider: str = "gemini"


class BatchAnalysisRequest(BaseModel):
    batch_id: str
    provider: str = "gemini"


class BatchChatMessage(BaseModel):
    batch_id: str
    message: str
    provider: str = "gemini"


# ──────────────────────────────────────────────────
# FastAPI App
# ──────────────────────────────────────────────────

app = FastAPI(title="Fraud Guard Pro", version="2.0")

app.mount("/static", StaticFiles(directory=str(BASE_DIR / "static")), name="static")


@app.get("/")
def index():
    return FileResponse(str(BASE_DIR / "static" / "index.html"))


@app.get("/models/")
def list_models():
    """Return available models for the frontend dropdown."""
    models = []
    for model_id, info in MODEL_REGISTRY.items():
        models.append({
            "id": model_id,
            "display_name": info["display_name"],
            "description": info.get("description", ""),
        })
    return {"models": models, "default": "random_forest_smote"}


@app.post("/predict/")
def predict(req: PredictRequest):
    try:
        model_id = req.model_id
        if model_id not in MODEL_REGISTRY:
            raise HTTPException(status_code=400, detail=f"Model '{model_id}' không tồn tại")

        ml_model = get_model(model_id)
        tx = req.transaction

        # Build DataFrame with correct column order
        columns = ["Time"] + [f"V{i}" for i in range(1, 29)] + ["Amount"]
        values = [tx.Time] + [getattr(tx, f"V{i}") for i in range(1, 29)] + [tx.Amount]
        df = pd.DataFrame([values], columns=columns)

        # Scale only Amount & Time (scaler was fit on these 2 columns only)
        if SCALER is None:
            raise HTTPException(status_code=500, detail="Scaler chưa được load")
        df[SCALE_COLS] = SCALER.transform(df[SCALE_COLS])

        # Predict using the full DataFrame (already scaled where needed)
        prediction = int(ml_model.predict(df)[0])

        # Get probability
        if hasattr(ml_model, "predict_proba"):
            proba = ml_model.predict_proba(df)[0]
            probs = np.abs(proba)
            total = np.sum(probs)
            if total > 0:
                normalized = probs / total
            else:
                normalized = [0.5, 0.5]
            fraud_prob = float(np.clip(normalized[1], 0.0, 1.0)) if len(normalized) > 1 else 0.0
        else:
            fraud_prob = 1.0 if prediction == 1 else 0.0

        # Determine verdict + confidence
        if fraud_prob >= 0.75:
            verdict = "Gian lận"
            confidence = fraud_prob  # % gian lận
        elif fraud_prob >= 0.35:
            verdict = "Nghi ngờ"
            confidence = fraud_prob  # % nghi ngờ gian lận
        else:
            verdict = "An toàn"
            confidence = 1.0 - fraud_prob  # % an toàn

        # Save to DB
        try:
            v_dict = {f"V{i}": getattr(tx, f"V{i}") for i in range(1, 29)}
            v_json = json.dumps(v_dict)
            conn = get_db()
            cur = conn.cursor()
            p = _ph()
            cur.execute(
                f"""INSERT INTO fraud_history
                   (model_used, time_val, amount, v_features, prediction, probability, verdict)
                   VALUES ({p}, {p}, {p}, {p}, {p}, {p}, {p})""",
                (MODEL_REGISTRY[model_id]["display_name"], tx.Time, tx.Amount,
                 v_json, prediction, confidence, verdict),
            )
            conn.commit()
            cur.close()
            conn.close()
        except Exception as db_err:
            logger.error("DB save failed: %s", db_err)

        return {
            "fraud_prediction": prediction,
            "probability": round(confidence, 4),
            "verdict": verdict,
            "model_used": MODEL_REGISTRY[model_id]["display_name"],
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error("Predict error: %s", e, exc_info=True)
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/history/")
def get_history():
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(
            "SELECT id, timestamp, amount, verdict, probability, model_used "
            "FROM fraud_history WHERE batch_id IS NULL ORDER BY timestamp DESC LIMIT 50"
        )
        rows = cur.fetchall()
        cur.close()
        conn.close()

        result = []
        for r in rows:
            ts = r[1]  # timestamp
            if isinstance(ts, str):
                try:
                    ts = datetime.strptime(ts, "%Y-%m-%d %H:%M:%S")
                except ValueError:
                    try:
                        ts = datetime.strptime(ts, "%Y-%m-%d %H:%M:%S.%f")
                    except ValueError:
                        pass
            if hasattr(ts, 'strftime'):
                # Convert UTC to Vietnam Time (UTC+7)
                ts = ts + timedelta(hours=7)
                ts = ts.strftime("%Y-%m-%d %H:%M:%S")
            result.append({
                "id": r[0],
                "timestamp": str(ts),
                "amount": r[2],
                "verdict": r[3],
                "probability": r[4],
                "model_used": r[5],
            })
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/history/{item_id}")
def get_history_item(item_id: int):
    try:
        conn = get_db()
        cur = conn.cursor()
        p = _ph()
        cur.execute(
            f"SELECT time_val, amount, v_features, prediction, probability, verdict, model_used "
            f"FROM fraud_history WHERE id = {p}",
            (item_id,),
        )
        r = cur.fetchone()
        cur.close()
        conn.close()

        if not r:
            raise HTTPException(status_code=404, detail="Not found")

        v_raw = r[2]  # v_features
        v_data = v_raw if isinstance(v_raw, dict) else json.loads(v_raw)
        return {
            "Time": r[0],
            "Amount": r[1],
            "V": v_data,
            "prediction": r[3],
            "probability": r[4],
            "verdict": r[5],
            "model_used": r[6],
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/chat/")
async def chat(msg: ChatMessage):
    context_str = ""
    if msg.context:
        context_str = f"\nNgữ cảnh giao dịch hiện tại: {json.dumps(msg.context, ensure_ascii=False)}"

    full_system = SYSTEM_PROMPT_CHAT + context_str
    response = await call_ai(msg.message, full_system, msg.provider)
    return {"response": response}


@app.post("/analyze/")
async def analyze(req: AnalysisRequest):
    status = "An Toàn" if req.prediction == 0 else "Gian lận"
    db_context = get_history_context()

    prompt = f"""BÁO CÁO PHÂN TÍCH FORENSIC TÀI CHÍNH
    ─────────────────────────────────
    MÔ HÌNH SỬ DỤNG: {req.model_used}
    TRẠNG THÁI: {status}
    ĐỘ TIN CẬY: {req.probability * 100:.2f}%
    DỮ LIỆU: {json.dumps(req.data, indent=2)}

    NGỮ CẢNH HỆ THỐNG:
    {db_context}

    YÊU CẦU PHÂN TÍCH CHI TIẾT:
    1. Đánh giá kỹ thuật các biến PCA (V1-V28), đặc biệt V14, V17, V12, V10, V4
    2. So sánh số tiền {req.data.get('Amount', 0)}$ với mức trung bình an toàn
    3. Đối chiếu với lịch sử giao dịch gần đây
    4. Kết luận chuyên gia chi tiết với thuật ngữ: PCA, Latent features, Anomaly Detection
    """

    response = await call_ai(prompt, SYSTEM_PROMPT_ANALYZE, req.provider)
    return {"analysis": response}


@app.post("/analyze-batch/")
async def analyze_batch(req: BatchAnalysisRequest):
    conn = get_db()
    cur = conn.cursor()
    p = _ph()
    
    cur.execute(f"SELECT customer_name, bank_name, model_used FROM fraud_history WHERE batch_id = {p} LIMIT 1", (req.batch_id,))
    meta = cur.fetchone()
    if not meta:
        cur.close()
        conn.close()
        raise HTTPException(status_code=404, detail="Batch not found")
    
    cur.execute(f"SELECT COUNT(*), SUM(CASE WHEN verdict='Gian lận' THEN 1 ELSE 0 END), SUM(CASE WHEN verdict='Nghi ngờ' THEN 1 ELSE 0 END) FROM fraud_history WHERE batch_id = {p}", (req.batch_id,))
    stats = cur.fetchone()
    
    cur.execute(f"SELECT time_val, amount, probability, verdict, tx_description FROM fraud_history WHERE batch_id = {p} AND verdict != 'An toàn' ORDER BY time_val ASC", (req.batch_id,))
    risky_rows = cur.fetchall()
    
    cur.close()
    conn.close()

    total = stats[0]
    fraud = stats[1] or 0
    suspect = stats[2] or 0
    risky_txs = [{"time_sec": r[0], "amount_vnd": r[1], "verdict": r[3], "confidence": f"{r[2]*100:.1f}%", "description": r[4]} for r in risky_rows]

    prompt = f"""BÁO CÁO PHÂN TÍCH FORENSIC LÔ GIAO DỊCH KHÁCH HÀNG
    ─────────────────────────────────
    KHÁCH HÀNG: {meta[0]} - NGÂN HÀNG: {meta[1]}
    MÔ HÌNH DỰ ĐOÁN: {meta[2]}
    TỔNG SỐ GIAO DỊCH: {total} | GIAN LẬN: {fraud} | NGHI NGỜ: {suspect}
    
    DANH SÁCH GIAO DỊCH RỦI RO CAO (Nghi ngờ / Gian lận):
    {json.dumps(risky_txs, ensure_ascii=False, indent=2)}

    YÊU CẦU: Hãy viết một đánh giá RẤT NGẮN GỌN (tối đa 3-4 câu) giải thích vì sao tài khoản này có rủi ro (dựa trên thời gian/số tiền/hành vi bất thường) và 1 câu kết luận/đề xuất hành động. Trình bày trơn tru dạng đoạn văn, KHÔNG chia đề mục hay gạch đầu dòng dài dòng.
    """

    response = await call_ai(prompt, SYSTEM_PROMPT_ANALYZE, req.provider)
    return {"analysis": response}


@app.post("/chat-batch/")
async def chat_batch(req: BatchChatMessage):
    conn = get_db()
    cur = conn.cursor()
    p = _ph()
    cur.execute(f"SELECT customer_name, bank_name, model_used FROM fraud_history WHERE batch_id = {p} LIMIT 1", (req.batch_id,))
    meta = cur.fetchone()
    
    if meta:
        cur.execute(f"SELECT COUNT(*), SUM(CASE WHEN verdict='Gian lận' THEN 1 ELSE 0 END), SUM(CASE WHEN verdict='Nghi ngờ' THEN 1 ELSE 0 END) FROM fraud_history WHERE batch_id = {p}", (req.batch_id,))
        stats = cur.fetchone()
        
        cur.execute(f"SELECT time_val, amount, verdict, tx_description FROM fraud_history WHERE batch_id = {p} ORDER BY time_val ASC", (req.batch_id,))
        all_txs = [{"time": r[0], "amount": r[1], "verdict": r[2], "desc": r[3]} for r in cur.fetchall()]
        
        context_str = f"\nNgữ cảnh: Đang trao đổi về Lô Giao Dịch của khách hàng: {meta[0]} ({meta[1]}). Tổng GD: {stats[0]}, Gian lận: {stats[1] or 0}, Nghi ngờ: {stats[2] or 0}.\nChi tiết toàn bộ GD:\n{json.dumps(all_txs, ensure_ascii=False)}"
    else:
        context_str = "\nNgữ cảnh: Không tìm thấy thông tin lô giao dịch."
    cur.close()
    conn.close()

    full_system = SYSTEM_PROMPT_CHAT + context_str
    response = await call_ai(req.message, full_system, req.provider)
    return {"response": response}


@app.delete("/history/{item_id}")
def delete_history_item(item_id: int):
    try:
        conn = get_db()
        cur = conn.cursor()
        p = _ph()
        cur.execute(f"DELETE FROM fraud_history WHERE id = {p}", (item_id,))
        deleted = cur.rowcount
        conn.commit()
        cur.close()
        conn.close()
        if deleted == 0:
            raise HTTPException(status_code=404, detail="Not found")
        return {"status": "success"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Delete error: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


# ──────────────────────────────────────────────────
# Batch Predict — Excel Import
# ──────────────────────────────────────────────────

REQUIRED_TECH_COLS = ["Time", "Amount"] + [f"V{i}" for i in range(1, 29)]
SAMPLE_DATA_DIR = BASE_DIR / "sample_data"


@app.post("/batch-predict/")
async def batch_predict(
    file: UploadFile = File(...),
    model_id: str = Form("random_forest_smote"),
):
    """Upload an Excel file, validate format, predict each row, save to DB."""

    # 1. Validate file type
    if not file.filename:
        raise HTTPException(400, "Không có file được chọn.")
    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if ext not in ("xlsx", "csv"):
        raise HTTPException(400, f"Định dạng '{ext}' không hỗ trợ. Chỉ chấp nhận .xlsx hoặc .csv")

    # 2. Validate model
    if model_id not in MODEL_REGISTRY:
        raise HTTPException(400, f"Mô hình '{model_id}' không tồn tại.")

    content = await file.read()

    # 3. Parse file
    try:
        if ext == "csv":
            df_tech = pd.read_csv(io.BytesIO(content))
            df_display = df_tech.copy()
            customer_name = "CSV Import"
            bank_name = ""
        else:
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
            # Check required sheet
            sheet_names = [s.lower().strip() for s in wb.sheetnames]
            tech_sheet = None
            display_sheet = None
            for s in wb.sheetnames:
                sl = s.lower().strip()
                if "ky thuat" in sl or "du lieu" in sl or "technical" in sl:
                    tech_sheet = s
                if "sao ke" in sl or "giao dich" in sl or "statement" in sl:
                    display_sheet = s

            if not tech_sheet:
                raise HTTPException(400,
                    "Không tìm thấy sheet 'Du lieu ky thuat'. "
                    "File Excel phải có sheet chứa dữ liệu kỹ thuật (Time, Amount, V1-V28).")

            # Read tech sheet
            ws_tech = wb[tech_sheet]
            tech_data = list(ws_tech.values)
            if len(tech_data) < 2:
                raise HTTPException(400, "Sheet dữ liệu kỹ thuật không có dữ liệu.")
            df_tech = pd.DataFrame(tech_data[1:], columns=tech_data[0])

            # Read display sheet for metadata + descriptions
            customer_name = ""
            bank_name = ""
            display_rows = []
            if display_sheet:
                ws_disp = wb[display_sheet]
                all_rows = list(ws_disp.values)
                # Extract metadata from first rows
                for row in all_rows[:4]:
                    if row and row[0]:
                        val = str(row[0])
                        if "khách hàng" in val.lower() or "kh:" in val.lower():
                            customer_name = val.split(":", 1)[-1].strip()
                        elif "ngân hàng" in val.lower() or "nh:" in val.lower():
                            bank_name = val.split(":", 1)[-1].strip()
                # Find header row (has "Ngày" or "STT")
                header_idx = None
                for i, row in enumerate(all_rows):
                    if row and any(str(c).lower().strip() in ("ngày gd", "stt", "ngày") for c in row if c):
                        header_idx = i
                        break
                if header_idx is not None and header_idx + 1 < len(all_rows):
                    display_rows = all_rows[header_idx + 1:]

            wb.close()

    except HTTPException:
        raise
    except Exception as e:
        logger.error("File parse error: %s", e)
        raise HTTPException(400, f"Lỗi đọc file: {str(e)}")

    # 4. Validate required columns
    missing = [c for c in REQUIRED_TECH_COLS if c not in df_tech.columns]
    if missing:
        raise HTTPException(400,
            f"Thiếu {len(missing)} cột bắt buộc: {', '.join(missing[:5])}... "
            f"File phải có các cột: Time, Amount, V1-V28")

    # 5. Clean data
    df_tech = df_tech[REQUIRED_TECH_COLS].apply(pd.to_numeric, errors="coerce").dropna()
    if len(df_tech) == 0:
        raise HTTPException(400, "Không có dữ liệu hợp lệ sau khi xử lý.")

    # 6. Run batch prediction
    ml_model = get_model(model_id)
    columns = ["Time"] + [f"V{i}" for i in range(1, 29)] + ["Amount"]
    df_predict = df_tech[columns].copy()

    if SCALER is None:
        raise HTTPException(500, "Scaler chưa được load.")
    df_predict[SCALE_COLS] = SCALER.transform(df_predict[SCALE_COLS])

    predictions = ml_model.predict(df_predict)
    if hasattr(ml_model, "predict_proba"):
        probas = ml_model.predict_proba(df_predict)
        fraud_probs = probas[:, 1] if probas.shape[1] > 1 else np.where(predictions == 1, 1.0, 0.0)
    else:
        fraud_probs = np.where(predictions == 1, 1.0, 0.0)

    # 7. Build results
    batch_id = str(uuid.uuid4())[:8]
    model_display = MODEL_REGISTRY[model_id]["display_name"]
    results = []

    for i in range(len(df_tech)):
        pred = int(predictions[i])
        prob = float(np.clip(fraud_probs[i], 0.0, 1.0))

        # Verdict + confidence (% tương ứng với kết luận)
        if prob >= 0.75:
            verdict = "Gian lận"
            confidence = prob
        elif prob >= 0.35:
            verdict = "Nghi ngờ"
            confidence = prob
        else:
            verdict = "An toàn"
            confidence = 1.0 - prob  # % an toàn

        time_val = float(df_tech.iloc[i]["Time"])
        amount_val = float(df_tech.iloc[i]["Amount"])
        v_dict = {f"V{j}": float(df_tech.iloc[i][f"V{j}"]) for j in range(1, 29)}

        # Get display info from sheet 1
        tx_desc = ""
        display_date = ""
        display_time = ""
        display_amount = ""
        if i < len(display_rows) and display_rows[i]:
            row = display_rows[i]
            display_date = str(row[0]) if row[0] else ""
            display_time = str(row[1]) if len(row) > 1 and row[1] else ""  # Giờ GD
            tx_desc = str(row[2]) if len(row) > 2 and row[2] else ""      # Nội dung
            display_amount = str(row[5]) if len(row) > 5 and row[5] else ""  # Số tiền VNĐ

        results.append({
            "row": i + 1,
            "date": display_date,
            "time": display_time,
            "description": tx_desc,
            "amount_display": display_amount,
            "amount": amount_val,
            "time_val": time_val,
            "v_features": v_dict,
            "prediction": pred,
            "probability": round(confidence, 4),
            "verdict": verdict,
        })

    # 8. Save to DB (non-blocking)
    try:
        conn = get_db()
        cur = conn.cursor()
        p = _ph()
        for r in results:
            cur.execute(
                f"""INSERT INTO fraud_history
                   (model_used, time_val, amount, v_features, prediction, probability, verdict,
                    batch_id, customer_name, bank_name, tx_description)
                   VALUES ({p},{p},{p},{p},{p},{p},{p},{p},{p},{p},{p})""",
                (model_display, r["time_val"], r["amount"], json.dumps(r["v_features"]),
                 r["prediction"], r["probability"], r["verdict"],
                 batch_id, customer_name, bank_name, r["description"]),
            )
        conn.commit()
        cur.close()
        conn.close()
    except Exception as e:
        logger.error("Batch DB save error: %s", e)

    # 8. Summary
    total = len(results)
    fraud_count = sum(1 for r in results if r["verdict"] == "Gian lận")
    suspect_count = sum(1 for r in results if r["verdict"] == "Nghi ngờ")
    safe_count = total - fraud_count - suspect_count

    return {
        "batch_id": batch_id,
        "customer_name": customer_name,
        "bank_name": bank_name,
        "model_used": model_display,
        "total": total,
        "safe": safe_count,
        "suspect": suspect_count,
        "fraud": fraud_count,
        "results": results,
    }


@app.get("/batch-history/")
def get_batch_history():
    """Get list of all batch imports."""
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            SELECT batch_id, customer_name, bank_name, model_used,
                   COUNT(*) as total,
                   SUM(CASE WHEN verdict = 'An toàn' THEN 1 ELSE 0 END) as safe_count,
                   SUM(CASE WHEN verdict = 'Gian lận' THEN 1 ELSE 0 END) as fraud_count,
                   MIN(timestamp) as imported_at
            FROM fraud_history
            WHERE batch_id IS NOT NULL AND batch_id != ''
            GROUP BY batch_id, customer_name, bank_name, model_used
            ORDER BY MIN(timestamp) DESC
            LIMIT 50
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()

        result = []
        for r in rows:
            ts = r[7]
            if isinstance(ts, str):
                try: ts = datetime.strptime(ts, "%Y-%m-%d %H:%M:%S")
                except ValueError: pass
            if hasattr(ts, 'strftime'):
                ts = ts + timedelta(hours=7)
                ts = ts.strftime("%Y-%m-%d %H:%M")
            result.append({
                "batch_id": r[0], "customer_name": r[1] or "", "bank_name": r[2] or "",
                "model_used": r[3] or "", "total": r[4], "safe": r[5], "fraud": r[6],
                "imported_at": str(ts),
            })
        return result
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/batch-history/{batch_id}")
def get_batch_detail(batch_id: str):
    """Get all transactions for a specific batch."""
    try:
        conn = get_db()
        cur = conn.cursor()
        p = _ph()
        cur.execute(
            f"""SELECT id, timestamp, amount, verdict, probability, model_used,
                       tx_description, customer_name, bank_name
                FROM fraud_history WHERE batch_id = {p}
                ORDER BY id ASC""",
            (batch_id,),
        )
        rows = cur.fetchall()
        cur.close()
        conn.close()

        if not rows:
            raise HTTPException(404, "Batch not found")

        result = []
        for r in rows:
            result.append({
                "id": r[0], "timestamp": str(r[1]), "amount": r[2],
                "verdict": r[3], "probability": r[4], "model_used": r[5],
                "description": r[6] or "", "customer_name": r[7] or "",
                "bank_name": r[8] or "",
            })
        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))


@app.delete("/batch-history/{batch_id}")
def delete_batch(batch_id: str):
    """Delete all transactions in a batch."""
    try:
        conn = get_db()
        cur = conn.cursor()
        p = _ph()
        cur.execute(f"DELETE FROM fraud_history WHERE batch_id = {p}", (batch_id,))
        deleted = cur.rowcount
        conn.commit()
        cur.close()
        conn.close()
        return {"status": "success", "deleted": deleted}
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/sample-files/")
def list_sample_files():
    """List available sample Excel files."""
    files = []
    if SAMPLE_DATA_DIR.exists():
        for f in sorted(SAMPLE_DATA_DIR.iterdir()):
            if f.suffix in (".xlsx", ".csv"):
                files.append({"name": f.name, "size": f.stat().st_size})
    return {"files": files}


@app.get("/sample-files/{filename}")
def download_sample_file(filename: str):
    """Download a sample Excel file."""
    path = SAMPLE_DATA_DIR / filename
    if not path.exists() or path.suffix not in (".xlsx", ".csv"):
        raise HTTPException(404, "File not found")
    return FileResponse(str(path), filename=filename,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ──────────────────────────────────────────────────
# Run
# ──────────────────────────────────────────────────

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
